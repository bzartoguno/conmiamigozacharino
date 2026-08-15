#!/usr/bin/env python3
"""
D&D Item Shop Sorter - Learned From "Copy of Current DND items"
===============================================================

This version does NOT rely mostly on generic guesses.

Every time it runs, it reads the workbook's:
    Copy of Current DND items

and uses the shop assignments already in that sheet as labeled examples of
how YOU organize your shops.

Then it sorts items in:
    DBCannonDndItems for programing

into:
    What shop it should go to

How it decides
--------------
1. Exact example match:
   If an item name already exists in the reference sheet, that shop wins.

2. Learned similarity:
   It builds a small local TF-IDF model from the 515-ish labeled examples.
   No AI service, web request, or external machine-learning package is used.

3. Shop guardrails:
   A few strong rules reflect patterns visible in your reference sheet:
     - Weapons -> Will's Weapons
     - Armor -> Supreme Smithy
     - Potions -> Pearl's Potions
     - Rings/jewelry -> Jewelry Guild
     - Clothing/fashion -> Iconic Dragonic
     - Transportation -> Navigation Guild
     - Death/revival -> Valhalla Mart
     - Corpses/undead -> Necromancy Insurance Company
     - Blood/HP trading -> N.M.E.
     - etc.

4. Review instead of pretending:
   Close/uncertain calls are still assigned a best guess, but are also added
   to a review sheet with:
     - 2nd and 3rd choices
     - confidence
     - reasons
     - closest examples from "Copy of Current DND items"

Important
---------
- It uses the LEFTMOST "What shop it should go to" column in the reference
  sheet as the real label column. The reference sheet currently has a second
  column with the same header; that duplicate is ignored.
- Existing manual shop assignments are preserved unless --overwrite is used.
- The original workbook is never overwritten.
- No internet is used.

Install:
    python -m pip install openpyxl

Run:
    python dnd_item_shop_sorter_learned.py

Dry run:
    python dnd_item_shop_sorter_learned.py --dry-run

Force recalculation of existing shop cells:
    python dnd_item_shop_sorter_learned.py --overwrite

Specific workbook:
    python dnd_item_shop_sorter_learned.py "E:\\DND infov2.xlsx"
"""

import argparse
import math
import os
import re
import shutil
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# WORKBOOK SETTINGS
# ---------------------------------------------------------------------------

TARGET_SHEET_EXACT = "DBCannonDndItems for programing"
TARGET_SHEET_CONTAINS = "DBCannonDndItems"
TRAINING_SHEET = "Copy of Current DND items"

ITEM_NAME_HEADER = "Item Name"
RARITY_HEADER = "Rarity"
ITEM_TYPE_HEADER = "Item Type"
DESCRIPTION_HEADER = "Description"
SHOP_HEADER = "What shop it should go to"

REVIEW_SHEET_NAME = "Shop Sorting Review v2"
OUTPUT_SUFFIX = " - SHOP SORTED LEARNED"
BACKUP_SUFFIX = " - BEFORE LEARNED SHOP SORT"

UNASSIGNED_VALUES = {
    "",
    "?",
    "???",
    "NA",
    "N/A",
    "NONE",
    "TBD",
}

# Set False if you ever want the program to write only shop names instead of
# the full labels from your reference sheet.
WRITE_FULL_SHOP_LABEL = True


# ---------------------------------------------------------------------------
# TEXT MODEL SETTINGS
# ---------------------------------------------------------------------------

STOPWORDS = set(
    """
    a an and are as at be been being but by can could did do does for from had
    has have he her hers him his how i if in into is it its may me more most my
    no not of on one only or our out over she so some than that the their them
    then there these they this to too up use used uses using was we were what
    when where which while who will with you your

    ability action advantage against any attack bonus creature creatures damage
    day effect feet ft gain gains gives hit item items make makes magical magic
    number once roll save saving spell target targets turn until wearing wield
    within
    """.split()
)


# ---------------------------------------------------------------------------
# SHOP GUARDRAILS
#
# These are intentionally narrower than the old sorter.
# They are based on the way the reference shop inventory is organized.
# Learned similarity still participates in every decision.
# ---------------------------------------------------------------------------

# (shop prefix, phrases, boost, scope)
# scope:
#   "name" = phrase must be in item name
#   "both" = phrase may be in name or description
CUE_RULES = [
    (
        "Bullets, Buffs, & Beyond",
        ["ammunition", "ammo", "arrow", "arrows", "bolt", "bolts",
         "bullet", "bullets", "quiver"],
        0.95,
        "both",
    ),
    (
        "Iconic Dragonic",
        ["cloak", "robe", "robes", "boots", "hat", "clothes", "clothing",
         "garment", "fashion", "earring", "earrings", "crown", "slippers",
         "gloves", "glove", "outfit"],
        0.72,
        "name",
    ),
    (
        "Jewelry Guild",
        ["gem", "gemstone", "amulet", "necklace", "brooch", "pendant",
         "jewel", "jewelry", "pearl", "talisman", "circlet"],
        0.70,
        "name",
    ),
    (
        "Navigation Guild",
        ["teleport", "transport", "travel", "portal", "flying", "fly",
         "flight", "compass", "ship", "boat", "broom", "carpet",
         "movement speed", "flying speed", "swimming speed"],
        0.65,
        "both",
    ),
    (
        "Robin's Ropes",
        ["rope", "grappling hook", "cord", "chain", "net", "lasso", "tether"],
        0.85,
        "name",
    ),
    (
        "Fairies of Flora",
        ["flower", "floral", "plant", "vine", "seed", "tree", "leaf",
         "leaves", "thorn", "root", "druid", "herb", "blossom"],
        0.72,
        "both",
    ),
    (
        "Silent Oath",
        ["assassin", "assassination", "stealth", "invisible",
         "invisibility", "shadow", "disguise", "spy", "silence"],
        0.62,
        "both",
    ),
    (
        "Changing Church",
        ["holy", "cleric", "divine", "blessing", "bless", "prayer",
         "celestial", "saint", "sacred", "holy symbol"],
        0.70,
        "both",
    ),
    (
        "O-Papies Oracle Readings",
        ["divination", "future", "fate", "oracle", "scry", "prophecy",
         "omen", "foresight", "fortune"],
        0.82,
        "both",
    ),
    (
        "Necromancy Insurance Company",
        ["corpse", "dead body", "undead", "necromancy", "zombie", "skeleton",
         "lich", "animate dead", "remains"],
        0.80,
        "both",
    ),
    (
        "Valhalla Mart",
        ["resurrection", "return to life", "restore to life",
         "death saving throw", "fallen hero", "deceased", "retired pc",
         "fallen pc", "afterlife"],
        0.76,
        "both",
    ),
    (
        "N.M.E.",
        ["blood", "hit points", "maximum hit points", "max hit points",
         "life force", "vampire", "sacrifice hit points"],
        0.58,
        "both",
    ),
    (
        "Paws, Claws, & Maws",
        ["familiar", "pet", "cub", "kitten", "pup", "egg", "mascot",
         "companion animal"],
        0.70,
        "both",
    ),
    (
        "Michael's Mount",
        ["mount", "steed", "saddle", "horseshoe", "ride", "riding"],
        0.82,
        "both",
    ),
    (
        "Golem Workshop",
        ["golem", "construct", "automaton", "mechanical servant"],
        0.76,
        "both",
    ),
    (
        "Jell Bell",
        ["slime", "ooze", "jelly", "gelatinous"],
        0.95,
        "name",
    ),
    (
        "Fizzy Tales",
        ["candy", "sweet", "sugar", "bubblegum", "gum", "chocolate",
         "mint", "marshmallow", "confection"],
        0.90,
        "both",
    ),
    (
        "Auntie Patty's Pies",
        ["pie", "tart", "food", "meal", "ration", "rations",
         "nourishment", "bread", "edible", "feast"],
        0.72,
        "both",
    ),
    (
        "Ye Old Donkey",
        ["coffee", "ale", "beer", "wine", "mead", "tankard", "drink",
         "drinking", "mug", "tavern"],
        0.85,
        "both",
    ),
    (
        "Book Bombs",
        ["book", "tome", "manual", "grimoire", "spellbook", "codex",
         "libram"],
        0.78,
        "name",
    ),
    (
        "Dungeon Crawler Guild",
        ["dungeon", "delving", "trap", "secret door", "cave", "piton",
         "underground"],
        0.74,
        "both",
    ),
    (
        "Sleuth University",
        ["investigation", "investigate", "detect", "clue", "forensic",
         "identify", "spyglass", "truth", "evidence", "tracking"],
        0.66,
        "both",
    ),
    (
        "Provision's Paradise",
        ["bag", "bottle", "chest", "container", "candle", "lantern", "jug",
         "storage", "bedroll", "box", "bucket"],
        0.48,
        "name",
    ),
    (
        "Ye Old Home Depot",
        ["tool", "nail", "door", "key", "lock", "ladder", "shovel", "saw",
         "chisel", "concrete", "lumber", "hinge"],
        0.68,
        "name",
    ),
    (
        "Comedy Gold",
        ["instrument", "music", "musical", "performance", "theatre",
         "theater", "dance", "bard", "mask", "costume", "stage"],
        0.70,
        "both",
    ),
    (
        "Goblin Market",
        ["explosive", "bomb", "grenade", "vehicle", "dynamite", "cannon",
         "chainsaw", "gunpowder", "rocket"],
        0.72,
        "both",
    ),
    (
        "Find a Friend",
        ["sentient", "intelligent item", "personality", "unique npc",
         "friend"],
        0.60,
        "both",
    ),
    (
        "Runestone Relay",
        ["message", "sending", "mail", "communicate", "communication",
         "telepathy", "letter", "correspondence"],
        0.78,
        "both",
    ),
    (
        "Blossom Hotel",
        ["hotel", "room", "board", "spa", "sleep", "rest", "pillow",
         "blanket", "lodging"],
        0.72,
        "both",
    ),
    (
        "The Piggy Bank",
        ["bank", "account", "coin", "currency", "gold", "money", "quest",
         "wager"],
        0.48,
        "both",
    ),
    (
        "Archives Guild",
        ["pawn", "sell item", "appraise", "appraisal", "antique",
         "collector"],
        0.62,
        "both",
    ),
    (
        "Labyrinthine Library",
        ["knowledge", "lore", "research", "scholar", "learning",
         "understanding", "memory"],
        0.58,
        "both",
    ),
    (
        "Make a Monster",
        ["summon a monster", "summon a creature", "summons a creature",
         "bodyguard", "guardian creature"],
        0.58,
        "both",
    ),
    (
        "Applegarth Guild",
        ["hire", "hireling", "servant", "retainer", "worker", "crew",
         "professional"],
        0.72,
        "both",
    ),
]


# ---------------------------------------------------------------------------
# TEXT HELPERS
# ---------------------------------------------------------------------------

def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalized(value) -> str:
    text = clean_text(value).lower()
    text = (
        text.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("–", "-")
        .replace("—", "-")
    )
    return re.sub(r"\s+", " ", text).strip()


def phrase_present(text: str, phrase: str) -> bool:
    """
    Match whole words, not substrings.

    This is important. For example:
      "hat" must NOT match "that"
      "rope" must NOT match "property"
      "omen" must NOT match "moment"
    """
    text = normalized(text)
    phrase = normalized(phrase)

    if not phrase:
        return False

    if " " in phrase:
        return phrase in text

    return re.search(
        rf"\b{re.escape(phrase)}\b",
        text,
    ) is not None


def tokenise(value) -> List[str]:
    text = normalized(value)
    words = re.findall(
        r"[a-z][a-z0-9']+",
        text,
    )

    tokens: List[str] = []

    for word in words:
        if word in STOPWORDS:
            continue

        if len(word) <= 2:
            continue

        if word.endswith("'s"):
            word = word[:-2]

        # Small, conservative stemmer. It just helps "travels" and "travel"
        # resemble each other. It deliberately avoids aggressive NLP.
        for suffix in ("ing", "ed", "es", "s"):
            if (
                len(word) > 6
                and word.endswith(suffix)
            ):
                word = word[:-len(suffix)]
                break

        if word and word not in STOPWORDS:
            tokens.append(word)

    return tokens


def weighted_document(
    item_name: str,
    item_type: str,
    description: str,
) -> List[str]:
    # Item name matters much more than an incidental word buried in a long
    # description.
    return (
        tokenise(item_name) * 4
        + tokenise(item_type) * 2
        + tokenise(description)
    )


def canonical_item_type(value: str) -> str:
    text = normalized(value)

    if text.startswith("weapon"):
        return "weapon"
    if text.startswith("armor"):
        return "armor"
    if text.startswith("potion"):
        return "potion"
    if text.startswith("ring"):
        return "ring"
    if text.startswith("scroll"):
        return "scroll"
    if text.startswith("wand"):
        return "wand"
    if text.startswith("staff"):
        return "staff"
    if text.startswith("rod"):
        return "rod"

    return text


def is_unassigned(value) -> bool:
    return normalized(value).upper() in UNASSIGNED_VALUES


# ---------------------------------------------------------------------------
# WORKBOOK HELPERS
# ---------------------------------------------------------------------------

def script_folder() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


def find_default_workbook() -> Optional[Path]:
    """
    Prefer an Excel workbook beside the script.
    """
    folders: List[Path] = []

    for folder in (
        script_folder(),
        Path.cwd(),
    ):
        if folder not in folders:
            folders.append(folder)

    preferred = (
        "DND infov2.xlsx",
        "DND info(2).xlsx",
        "DND info.xlsx",
        "DND info(1).xlsx",
    )

    for folder in folders:
        for filename in preferred:
            candidate = folder / filename

            if candidate.is_file():
                return candidate

    candidates: List[Path] = []

    for folder in folders:
        for candidate in folder.glob("DND info*.xlsx"):
            upper = candidate.name.upper()

            if candidate.name.startswith("~$"):
                continue

            if "SHOP SORTED" in upper:
                continue

            if "BEFORE" in upper:
                continue

            if candidate not in candidates:
                candidates.append(candidate)

    if len(candidates) == 1:
        return candidates[0]

    return None


def find_target_sheet(wb):
    if TARGET_SHEET_EXACT in wb.sheetnames:
        return wb[TARGET_SHEET_EXACT]

    for name in wb.sheetnames:
        if TARGET_SHEET_CONTAINS.lower() in name.lower():
            return wb[name]

    return None


def header_positions(ws) -> Dict[str, List[int]]:
    """
    Return header -> list of columns.

    Lists matter because the reference sheet currently has TWO columns named
    "What shop it should go to". We intentionally use the leftmost one.
    """
    result: Dict[str, List[int]] = defaultdict(list)

    for cell in ws[1]:
        if cell.value is None:
            continue

        result[
            clean_text(cell.value)
        ].append(cell.column)

    return result


def first_header_column(
    positions: Dict[str, List[int]],
    header: str,
) -> Optional[int]:
    cols = positions.get(header, [])
    return min(cols) if cols else None


# ---------------------------------------------------------------------------
# TRAINING MODEL
# ---------------------------------------------------------------------------

class TrainingExample:
    def __init__(
        self,
        row_number: int,
        name: str,
        rarity: str,
        item_type: str,
        description: str,
        shop: str,
    ):
        self.row_number = row_number
        self.name = name
        self.rarity = rarity
        self.item_type = item_type
        self.description = description
        self.shop = shop
        self.tokens = weighted_document(
            name,
            item_type,
            description,
        )
        self.vector: Dict[str, float] = {}


class ShopModel:
    def __init__(
        self,
        training_examples: List[TrainingExample],
    ):
        self.examples = training_examples
        self.shops = sorted(
            {
                example.shop
                for example in training_examples
                if example.shop
            }
        )

        self.shop_by_prefix: Dict[str, str] = {}
        for shop in self.shops:
            short = shop.split(" (", 1)[0]
            self.shop_by_prefix[short] = shop

        self.exact_name_to_shop: Dict[str, str] = {}
        for example in training_examples:
            key = normalized(example.name)

            if (
                key
                and key not in self.exact_name_to_shop
            ):
                self.exact_name_to_shop[key] = example.shop

        self.examples_by_shop: Dict[
            str,
            List[TrainingExample],
        ] = defaultdict(list)

        for example in training_examples:
            self.examples_by_shop[
                example.shop
            ].append(example)

        self.idf: Dict[str, float] = {}
        self.centroids: Dict[
            str,
            Dict[str, float],
        ] = {}
        self.label_vectors: Dict[
            str,
            Dict[str, float],
        ] = {}

        self._fit()

    def _fit(self) -> None:
        document_frequency: Counter = Counter()

        for example in self.examples:
            document_frequency.update(
                set(example.tokens)
            )

        document_count = max(
            1,
            len(self.examples),
        )

        self.idf = {
            token: (
                math.log(
                    (document_count + 1)
                    / (frequency + 1)
                )
                + 1.0
            )
            for token, frequency
            in document_frequency.items()
        }

        for example in self.examples:
            example.vector = self.vectorise(
                example.tokens
            )

        for shop, examples in self.examples_by_shop.items():
            accumulator: Dict[
                str,
                float,
            ] = defaultdict(float)

            for example in examples:
                for token, value in example.vector.items():
                    accumulator[token] += value

            count = max(
                1,
                len(examples),
            )

            centroid = {
                token: value / count
                for token, value
                in accumulator.items()
            }

            self.centroids[shop] = (
                self.normalise_vector(
                    centroid
                )
            )

            self.label_vectors[shop] = (
                self.vectorise(
                    tokenise(shop) * 3
                )
            )

    def vectorise(
        self,
        tokens: List[str],
    ) -> Dict[str, float]:
        counts = Counter(tokens)

        vector = {}

        for token, count in counts.items():
            idf = self.idf.get(
                token,
                1.0,
            )

            vector[token] = (
                (1.0 + math.log(count))
                * idf
            )

        return self.normalise_vector(
            vector
        )

    @staticmethod
    def normalise_vector(
        vector: Dict[str, float],
    ) -> Dict[str, float]:
        magnitude = math.sqrt(
            sum(
                value * value
                for value in vector.values()
            )
        )

        if magnitude <= 0:
            return {}

        return {
            token: value / magnitude
            for token, value
            in vector.items()
        }

    @staticmethod
    def cosine(
        left: Dict[str, float],
        right: Dict[str, float],
    ) -> float:
        if len(left) > len(right):
            left, right = right, left

        return sum(
            value
            * right.get(token, 0.0)
            for token, value
            in left.items()
        )

    def resolve_shop_prefix(
        self,
        prefix: str,
    ) -> Optional[str]:
        """
        Prefix matching makes the program tolerate the Navigation Guild label
        in the workbook currently missing its final ')'.
        """
        exact = self.shop_by_prefix.get(
            prefix
        )

        if exact:
            return exact

        for shop in self.shops:
            if normalized(shop).startswith(
                normalized(prefix)
            ):
                return shop

        return None

    def semantic_scores(
        self,
        item_name: str,
        item_type: str,
        description: str,
    ) -> Tuple[
        Dict[str, float],
        Dict[str, List[Tuple[float, TrainingExample]]],
    ]:
        target_vector = self.vectorise(
            weighted_document(
                item_name,
                item_type,
                description,
            )
        )

        scores: Dict[str, float] = {}
        nearest: Dict[
            str,
            List[Tuple[float, TrainingExample]],
        ] = {}

        for shop, examples in self.examples_by_shop.items():
            similarities = []

            for example in examples:
                similarity = self.cosine(
                    target_vector,
                    example.vector,
                )

                similarities.append(
                    (
                        similarity,
                        example,
                    )
                )

            similarities.sort(
                key=lambda pair: pair[0],
                reverse=True,
            )

            nearest[shop] = similarities[:3]

            top_values = [
                pair[0]
                for pair in similarities[:3]
            ]

            if top_values:
                top_average = (
                    sum(top_values)
                    / len(top_values)
                )
            else:
                top_average = 0.0

            centroid_similarity = self.cosine(
                target_vector,
                self.centroids.get(
                    shop,
                    {},
                ),
            )

            label_similarity = self.cosine(
                target_vector,
                self.label_vectors.get(
                    shop,
                    {},
                ),
            )

            # Deliberately modest. The examples guide ambiguous cases, while
            # strong object/store patterns are handled by the guardrails.
            scores[shop] = (
                0.20 * top_average
                + 0.10 * centroid_similarity
                + 0.05 * label_similarity
            )

        return scores, nearest


def read_training_examples(
    wb,
) -> List[TrainingExample]:
    if TRAINING_SHEET not in wb.sheetnames:
        raise ValueError(
            f'Training sheet "{TRAINING_SHEET}" '
            "was not found."
        )

    ws = wb[TRAINING_SHEET]
    positions = header_positions(ws)

    name_col = first_header_column(
        positions,
        ITEM_NAME_HEADER,
    )
    rarity_col = first_header_column(
        positions,
        RARITY_HEADER,
    )
    type_col = first_header_column(
        positions,
        ITEM_TYPE_HEADER,
    )
    description_col = first_header_column(
        positions,
        DESCRIPTION_HEADER,
    )

    # IMPORTANT: choose the LEFTMOST duplicate shop column.
    shop_cols = positions.get(
        SHOP_HEADER,
        [],
    )

    if not shop_cols:
        raise ValueError(
            f'Training sheet is missing "{SHOP_HEADER}".'
        )

    shop_col = min(shop_cols)

    for header, col in (
        (ITEM_NAME_HEADER, name_col),
        (RARITY_HEADER, rarity_col),
        (ITEM_TYPE_HEADER, type_col),
        (DESCRIPTION_HEADER, description_col),
    ):
        if col is None:
            raise ValueError(
                f'Training sheet is missing "{header}".'
            )

    examples: List[TrainingExample] = []

    for row in range(
        2,
        ws.max_row + 1,
    ):
        name = clean_text(
            ws.cell(
                row=row,
                column=name_col,
            ).value
        )

        shop = clean_text(
            ws.cell(
                row=row,
                column=shop_col,
            ).value
        )

        if not name or not shop:
            continue

        examples.append(
            TrainingExample(
                row_number=row,
                name=name,
                rarity=clean_text(
                    ws.cell(
                        row=row,
                        column=rarity_col,
                    ).value
                ),
                item_type=clean_text(
                    ws.cell(
                        row=row,
                        column=type_col,
                    ).value
                ),
                description=clean_text(
                    ws.cell(
                        row=row,
                        column=description_col,
                    ).value
                ),
                shop=shop,
            )
        )

    if not examples:
        raise ValueError(
            "No labeled training examples were found."
        )

    return examples


# ---------------------------------------------------------------------------
# CLASSIFIER
# ---------------------------------------------------------------------------

def add_score(
    scores: Dict[str, float],
    reasons: Dict[str, List[str]],
    model: ShopModel,
    shop_prefix: str,
    amount: float,
    reason: str,
) -> None:
    shop = model.resolve_shop_prefix(
        shop_prefix
    )

    if shop is None:
        return

    scores[shop] = (
        scores.get(shop, 0.0)
        + amount
    )

    if reason not in reasons[shop]:
        reasons[shop].append(reason)


def classify_item(
    model: ShopModel,
    item_name: str,
    rarity: str,
    item_type: str,
    description: str,
) -> Tuple[
    str,
    float,
    float,
    str,
    List[Tuple[str, float]],
    List[str],
    List[str],
]:
    """
    Return:
        chosen_shop
        chosen_score
        margin
        confidence
        top_three [(shop, score), ...]
        reasons_for_chosen_shop
        nearest_example_strings
    """

    exact_shop = model.exact_name_to_shop.get(
        normalized(item_name)
    )

    if exact_shop:
        return (
            exact_shop,
            10.0,
            10.0,
            "High",
            [(exact_shop, 10.0)],
            [
                'Exact item-name match in '
                f'"{TRAINING_SHEET}"'
            ],
            [
                f"{item_name} -> {exact_shop}"
            ],
        )

    scores, nearest = model.semantic_scores(
        item_name,
        item_type,
        description,
    )

    reasons: Dict[
        str,
        List[str],
    ] = defaultdict(list)

    canonical_type = canonical_item_type(
        item_type
    )

    rarity_text = normalized(
        rarity
    )

    name_text = normalized(
        item_name
    )

    combined_text = (
        name_text
        + " "
        + normalized(description)
    )

    # ---------------------------------------------------------------
    # Strong store patterns visible in the reference inventory.
    # ---------------------------------------------------------------

    if canonical_type == "weapon":
        add_score(
            scores,
            reasons,
            model,
            "Will's Weapons",
            0.75,
            "Item Type is Weapon",
        )

    elif canonical_type == "armor":
        add_score(
            scores,
            reasons,
            model,
            "Supreme Smithy",
            0.80,
            "Item Type is Armor",
        )

    elif canonical_type == "potion":
        # Pearl's has the clear potion inventory in the reference sheet.
        add_score(
            scores,
            reasons,
            model,
            "Pearl's Potions",
            0.90,
            "Item Type is Potion",
        )

        # The store label says "Better Potions". The reference sheet has only
        # one Jazz example, so this remains a smaller hint rather than a hard
        # rule. High-tier potions get a little extra consideration.
        add_score(
            scores,
            reasons,
            model,
            "Jazz's Portable Potions",
            (
                0.34
                if rarity_text
                in {
                    "very rare",
                    "legendary",
                    "artifact",
                }
                else 0.10
            ),
            "Jazz is labeled Better Potions",
        )

    elif canonical_type == "ring":
        add_score(
            scores,
            reasons,
            model,
            "Jewelry Guild",
            0.72,
            "Item Type is Ring",
        )

    elif canonical_type == "scroll":
        add_score(
            scores,
            reasons,
            model,
            "Book Bombs",
            0.55,
            "Item Type is Scroll",
        )

        add_score(
            scores,
            reasons,
            model,
            "Evan's Enchanting Emporium",
            0.20,
            "Scroll is also an enchanting/magic item",
        )

    elif canonical_type in {
        "wand",
        "staff",
        "rod",
    }:
        add_score(
            scores,
            reasons,
            model,
            "Evan's Enchanting Emporium",
            0.62,
            f"Item Type is {canonical_type.title()}",
        )

    # Wondrous armor-shaped objects are still likely to belong with armor.
    armor_words = (
        "helm",
        "helmet",
        "shield",
        "breastplate",
        "armor",
        "armour",
        "plate armor",
        "mail",
    )

    if any(
        phrase_present(
            name_text,
            phrase,
        )
        for phrase in armor_words
    ):
        add_score(
            scores,
            reasons,
            model,
            "Supreme Smithy",
            0.65,
            "Name is armor/shield/helm shaped",
        )

    # Wondrous weapon-shaped objects still lean toward Will's.
    weapon_words = (
        "sword",
        "blade",
        "axe",
        "dagger",
        "mace",
        "hammer",
        "spear",
        "bow",
        "crossbow",
        "staff",
        "whip",
        "javelin",
        "glaive",
        "halberd",
        "trident",
    )

    if (
        any(
            phrase_present(
                name_text,
                phrase,
            )
            for phrase in weapon_words
        )
        and not phrase_present(
            combined_text,
            "ammunition",
        )
        and not phrase_present(
            combined_text,
            "ammo",
        )
    ):
        add_score(
            scores,
            reasons,
            model,
            "Will's Weapons",
            0.45,
            "Name is weapon-shaped",
        )

    # Clothing descriptions can identify fashion even when the item has a
    # fantasy name such as Glamerweave.
    fashion_description_words = (
        "clothing",
        "garment",
        "outfit",
        "robe",
        "cloak",
        "wear as clothing",
    )

    if any(
        phrase_present(
            normalized(description),
            phrase,
        )
        for phrase in fashion_description_words
    ):
        add_score(
            scores,
            reasons,
            model,
            "Iconic Dragonic",
            0.32,
            "Description identifies clothing/fashion",
        )

    # ---------------------------------------------------------------
    # Thematic cues.
    # ---------------------------------------------------------------

    for (
        shop_prefix,
        phrases,
        boost,
        scope,
    ) in CUE_RULES:
        search_text = (
            name_text
            if scope == "name"
            else combined_text
        )

        matches = [
            phrase
            for phrase in phrases
            if phrase_present(
                search_text,
                phrase,
            )
        ]

        if not matches:
            continue

        name_matches = [
            phrase
            for phrase in matches
            if phrase_present(
                name_text,
                phrase,
            )
        ]

        amount = (
            boost
            + 0.10
            * min(
                2,
                max(
                    0,
                    len(matches) - 1,
                ),
            )
            + (
                0.16
                if name_matches
                else 0.0
            )
        )

        add_score(
            scores,
            reasons,
            model,
            shop_prefix,
            amount,
            "Theme match: "
            + ", ".join(
                matches[:3]
            ),
        )

    ranked = sorted(
        scores.items(),
        key=lambda pair: (
            pair[1],
            pair[0],
        ),
        reverse=True,
    )

    if not ranked:
        raise RuntimeError(
            f'No shop candidates were produced for "{item_name}".'
        )

    chosen_shop, chosen_score = ranked[0]

    second_score = (
        ranked[1][1]
        if len(ranked) > 1
        else 0.0
    )

    margin = (
        chosen_score
        - second_score
    )

    # Confidence deliberately leans toward review.
    if (
        chosen_score >= 0.75
        and margin >= 0.28
    ):
        confidence = "High"

    elif (
        chosen_score >= 0.42
        and margin >= 0.14
    ):
        confidence = "Medium"

    else:
        confidence = "Review"

    top_three = [
        (
            shop,
            round(score, 3),
        )
        for shop, score
        in ranked[:3]
    ]

    chosen_reasons = list(
        reasons.get(
            chosen_shop,
            [],
        )
    )

    if not chosen_reasons:
        chosen_reasons.append(
            "Closest fit to your labeled reference examples"
        )
    else:
        chosen_reasons.append(
            "Reference-example similarity also included"
        )

    nearest_examples: List[str] = []

    for similarity, example in nearest.get(
        chosen_shop,
        [],
    )[:3]:
        nearest_examples.append(
            f'{example.name} '
            f'(reference row {example.row_number}, '
            f'similarity {similarity:.3f})'
        )

    return (
        chosen_shop,
        round(chosen_score, 3),
        round(margin, 3),
        confidence,
        top_three,
        chosen_reasons,
        nearest_examples,
    )


# ---------------------------------------------------------------------------
# REVIEW SHEET
# ---------------------------------------------------------------------------

def format_review_sheet(ws) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    widths = {
        "A": 10,
        "B": 34,
        "C": 16,
        "D": 20,
        "E": 58,
        "F": 12,
        "G": 12,
        "H": 14,
        "I": 58,
        "J": 12,
        "K": 58,
        "L": 12,
        "M": 68,
        "N": 68,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    for row in ws.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    high_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    medium_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    review_fill = PatternFill(
        fill_type="solid",
        fgColor="FCE4D6",
    )

    for row in range(
        2,
        ws.max_row + 1,
    ):
        confidence = clean_text(
            ws.cell(
                row=row,
                column=8,
            ).value
        )

        if confidence == "High":
            fill = high_fill
        elif confidence == "Medium":
            fill = medium_fill
        else:
            fill = review_fill

        ws.cell(
            row=row,
            column=8,
        ).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# FILE SAFETY
# ---------------------------------------------------------------------------

def make_backup(
    input_path: Path,
) -> Path:
    backup = input_path.with_name(
        input_path.stem
        + BACKUP_SUFFIX
        + input_path.suffix
    )

    if not backup.exists():
        shutil.copy2(
            input_path,
            backup,
        )

    return backup


def output_path_for(
    input_path: Path,
) -> Path:
    return input_path.with_name(
        input_path.stem
        + OUTPUT_SUFFIX
        + input_path.suffix
    )


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Sort D&D items into shops using the labeled "
            '"Copy of Current DND items" sheet as training examples.'
        )
    )

    parser.add_argument(
        "workbook",
        nargs="?",
        default=None,
        help=(
            "Workbook path. If omitted, look beside the script."
        ),
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help=(
            "Analyze without writing or saving a workbook."
        ),
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "Re-sort rows that already contain a manual shop assignment."
        ),
    )

    args = parser.parse_args()

    if args.workbook:
        input_path = Path(
            args.workbook
        ).expanduser().resolve()
    else:
        found = find_default_workbook()

        if found is None:
            print()
            print(
                "ERROR: Could not identify the DND workbook."
            )
            print(
                "Put this script beside DND infov2.xlsx "
                "or pass the workbook path explicitly."
            )
            return 1

        input_path = found.resolve()

    if not input_path.is_file():
        print()
        print(
            "ERROR: Workbook not found:"
        )
        print(input_path)
        return 1

    print()
    print("=" * 78)
    print("D&D SHOP SORTER - LEARNED FROM YOUR CURRENT SHOP INVENTORY")
    print("=" * 78)
    print("Input:   ", input_path)
    print("Training:", TRAINING_SHEET)
    print("Target:  ", TARGET_SHEET_EXACT)
    print(
        "Mode:    ",
        "DRY RUN"
        if args.dry_run
        else "WRITE SHOP ASSIGNMENTS",
    )
    print(
        "Manual assignments:",
        "OVERWRITE"
        if args.overwrite
        else "PRESERVE",
    )
    print("Internet: NONE")
    print()

    try:
        wb = load_workbook(
            input_path
        )
    except Exception as error:
        print(
            "ERROR: Could not open workbook:"
        )
        print(error)
        return 1

    try:
        training_examples = (
            read_training_examples(
                wb
            )
        )
    except Exception as error:
        print(
            "ERROR: Could not build training set:"
        )
        print(error)
        return 1

    model = ShopModel(
        training_examples
    )

    print(
        "Training examples loaded:",
        len(training_examples),
    )
    print(
        "Shops learned:",
        len(model.shops),
    )
    print()

    target_ws = find_target_sheet(
        wb
    )

    if target_ws is None:
        print(
            "ERROR: Could not find the DBCannonDndItems sheet."
        )
        print("Available sheets:")

        for name in wb.sheetnames:
            print(" -", name)

        return 1

    positions = header_positions(
        target_ws
    )

    required_headers = (
        ITEM_NAME_HEADER,
        RARITY_HEADER,
        ITEM_TYPE_HEADER,
        DESCRIPTION_HEADER,
        SHOP_HEADER,
    )

    target_columns: Dict[
        str,
        int,
    ] = {}

    for header in required_headers:
        column = first_header_column(
            positions,
            header,
        )

        if column is None:
            print(
                f'ERROR: Target sheet is missing "{header}".'
            )
            return 1

        target_columns[header] = column

    name_col = target_columns[
        ITEM_NAME_HEADER
    ]
    rarity_col = target_columns[
        RARITY_HEADER
    ]
    type_col = target_columns[
        ITEM_TYPE_HEADER
    ]
    description_col = target_columns[
        DESCRIPTION_HEADER
    ]
    shop_col = target_columns[
        SHOP_HEADER
    ]

    print(
        "Shop output column:",
        get_column_letter(
            shop_col
        ),
        f'("{SHOP_HEADER}")',
    )
    print()

    total_items = 0
    assigned = 0
    preserved = 0
    confidence_counts = Counter()
    shop_counts = Counter()
    review_rows: List[List] = []

    for excel_row in range(
        2,
        target_ws.max_row + 1,
    ):
        item_name = clean_text(
            target_ws.cell(
                row=excel_row,
                column=name_col,
            ).value
        )

        if not item_name:
            continue

        total_items += 1

        current_shop = target_ws.cell(
            row=excel_row,
            column=shop_col,
        ).value

        if (
            not args.overwrite
            and not is_unassigned(
                current_shop
            )
        ):
            preserved += 1
            shop_counts[
                clean_text(
                    current_shop
                )
            ] += 1
            continue

        rarity = clean_text(
            target_ws.cell(
                row=excel_row,
                column=rarity_col,
            ).value
        )

        item_type = clean_text(
            target_ws.cell(
                row=excel_row,
                column=type_col,
            ).value
        )

        description = clean_text(
            target_ws.cell(
                row=excel_row,
                column=description_col,
            ).value
        )

        (
            chosen_shop,
            chosen_score,
            margin,
            confidence,
            top_three,
            reasons,
            nearest_examples,
        ) = classify_item(
            model=model,
            item_name=item_name,
            rarity=rarity,
            item_type=item_type,
            description=description,
        )

        if not args.dry_run:
            cell = target_ws.cell(
                row=excel_row,
                column=shop_col,
            )

            cell.value = (
                chosen_shop
                if WRITE_FULL_SHOP_LABEL
                else chosen_shop.split(
                    " (",
                    1,
                )[0]
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        assigned += 1
        confidence_counts[
            confidence
        ] += 1
        shop_counts[
            chosen_shop
        ] += 1

        # Only non-high-confidence rows need manual attention.
        if confidence != "High":
            candidates = list(
                top_three
            )

            while len(candidates) < 3:
                candidates.append(
                    ("", 0.0)
                )

            review_rows.append(
                [
                    excel_row,
                    item_name,
                    rarity,
                    item_type,
                    candidates[0][0],
                    candidates[0][1],
                    margin,
                    confidence,
                    candidates[1][0],
                    candidates[1][1],
                    candidates[2][0],
                    candidates[2][1],
                    "; ".join(
                        reasons
                    ),
                    "\n".join(
                        nearest_examples
                    ),
                ]
            )

    print(
        "Items found:             ",
        total_items,
    )
    print(
        "Assignments generated:   ",
        assigned,
    )
    print(
        "Manual assignments kept: ",
        preserved,
    )
    print()
    print(
        "High confidence:          ",
        confidence_counts[
            "High"
        ],
    )
    print(
        "Medium confidence:        ",
        confidence_counts[
            "Medium"
        ],
    )
    print(
        "Needs review:             ",
        confidence_counts[
            "Review"
        ],
    )
    print()

    print(
        "SHOP DISTRIBUTION"
    )
    print(
        "-" * 78
    )

    for shop, count in sorted(
        shop_counts.items(),
        key=lambda pair: (
            -pair[1],
            pair[0],
        ),
    ):
        print(
            f"{count:>4}  {shop}"
        )

    if args.dry_run:
        print()
        print(
            "Dry run complete. "
            "No workbook was changed or saved."
        )
        return 0

    if REVIEW_SHEET_NAME in wb.sheetnames:
        del wb[
            REVIEW_SHEET_NAME
        ]

    review_ws = wb.create_sheet(
        REVIEW_SHEET_NAME
    )

    review_ws.append(
        [
            "Excel Row",
            "Item Name",
            "Rarity",
            "Item Type",
            "Suggested Shop",
            "Score",
            "Lead Over 2nd",
            "Confidence",
            "2nd Choice",
            "2nd Score",
            "3rd Choice",
            "3rd Score",
            "Why It Chose This",
            "Closest Examples From Copy of Current DND items",
        ]
    )

    for row in review_rows:
        review_ws.append(row)

    format_review_sheet(
        review_ws
    )

    target_ws.column_dimensions[
        get_column_letter(
            shop_col
        )
    ].width = 60

    for row in range(
        2,
        target_ws.max_row + 1,
    ):
        target_ws.cell(
            row=row,
            column=shop_col,
        ).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

    backup = make_backup(
        input_path
    )

    output_path = output_path_for(
        input_path
    )

    wb.save(
        output_path
    )

    print()
    print("=" * 78)
    print("DONE")
    print("=" * 78)
    print(
        "Original:",
        input_path,
    )
    print()
    print(
        "Safety backup:",
        backup,
    )
    print()
    print(
        "Sorted workbook:",
        output_path,
    )
    print()
    print(
        f'Review sheet: "{REVIEW_SHEET_NAME}" '
        f"({len(review_rows)} rows)"
    )
    print()
    print(
        "The next time you run this program, edits you make in "
        '"Copy of Current DND items" automatically become part of '
        "the training examples."
    )
    print(
        "Manual shop choices in the target sheet are also preserved "
        "unless you use --overwrite."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
