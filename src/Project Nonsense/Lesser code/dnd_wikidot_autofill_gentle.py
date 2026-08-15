#!/usr/bin/env python3
"""
D&D Wikidot Magic Item Autofiller
=================================

Tailored to the workbook:
    DND info.xlsx

Target sheet:
    DBCannonDndItems for programing

What it does
------------
1. Opens the D&D 5e Wikidot Magic Items master page.
2. Collects item-name -> item-page hyperlinks automatically.
3. Opens your workbook.
4. Leaves every existing description alone.
5. For rows where Description is blank, finds the matching Wikidot item page.
6. Extracts the item description.
7. Writes it into the Description column.
8. Adds a "Wikidot URL" column for traceability.
9. Saves progress repeatedly so an interruption does not lose much work.
10. Writes unmatched / failed items to a text report.

It does NOT require you to manually add hyperlinks to the spreadsheet first.

Install dependencies:
    python3 -m pip install requests beautifulsoup4 openpyxl

Normal gentle run (defaults to at most 50 missing descriptions):
    python3 dnd_wikidot_autofill_gentle.py

Dry run (match/check only, no descriptions downloaded or workbook saved):
    python3 dnd_wikidot_autofill.py --dry-run

Process only the next 10 missing items:
    python3 dnd_wikidot_autofill_gentle.py --limit 10

Process up to the default 50-item cap:
    python3 dnd_wikidot_autofill_gentle.py

Use a different workbook:
    python3 dnd_wikidot_autofill.py "/path/to/DND info.xlsx"
"""

import argparse
import html
import random
import os
import re
import shutil
import sys
import time
import unicodedata
from copy import copy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Optional, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


MASTER_URL = "https://dnd5e.wikidot.com/wondrous-items"
DEFAULT_WORKBOOK = "DND info.xlsx"
TARGET_SHEET = "DBCannonDndItems for programing"

# =====================
# GENTLE REQUEST SETTINGS
# =====================
#
# These settings are for load smoothing and reducing repeated traffic.
# They are not intended to bypass rate limits or hide automation.

MIN_REQUEST_DELAY_SECONDS = 5.0
MAX_REQUEST_DELAY_SECONDS = 10.0

# Unless --limit is supplied, process at most this many missing descriptions
# in one run.
DEFAULT_MAX_ITEMS_PER_RUN = 50

REQUEST_TIMEOUT_SECONDS = 30
SAVE_EVERY_N_ITEMS = 5

# Retry/backoff behavior for temporary server trouble.
MAX_RETRIES_PER_PAGE = 3
BACKOFF_SECONDS = (30, 60, 120)

# If Wikidot explicitly rate-limits us, stop the run rather than hammering it.
STOP_ON_HTTP_429 = True

# Stop after repeated 403/5xx responses from one page.
STOP_ON_REPEATED_SERVER_ERRORS = True

# Cache downloaded item HTML beside the workbook so already-fetched pages
# do not need to be requested again on later runs.
CACHE_FOLDER_NAME = "dnd_wikidot_cache"

USER_AGENT = (
    "Mozilla/5.0 (compatible; Personal-DND-Spreadsheet-Helper/1.0; "
    "+https://dnd5e.wikidot.com/)"
)

# Only use a fuzzy match when it is extremely close.
# Exact normalized matching is always preferred.
FUZZY_MATCH_THRESHOLD = 0.94

DESCRIPTION_HEADER = "Description"
ITEM_NAME_HEADER = "Item Name"
URL_HEADER = "Wikidot URL"


def normalize_name(value: str) -> str:
    """
    Normalize names so harmless punctuation/capitalization differences match.

    Examples:
      "House Of Cards" == "house of cards"
      curly apostrophe == straight apostrophe
      extra whitespace is ignored
    """
    if value is None:
        return ""

    text = unicodedata.normalize("NFKC", str(value))
    text = (
        text.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("–", "-")
        .replace("—", "-")
        .strip()
        .lower()
    )

    # Treat common punctuation as spacing, then collapse whitespace.
    text = re.sub(r"[:;,()\[\]{}]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    return text


class StopGentleRun(Exception):
    """Raised when the server signals that the run should stop politely."""


def gentle_delay() -> float:
    """
    Sleep for a randomized interval to spread requests out over time.

    The randomness is only for load smoothing; it is not used to evade
    detection or rate limits.
    """
    delay = random.uniform(
        MIN_REQUEST_DELAY_SECONDS,
        MAX_REQUEST_DELAY_SECONDS,
    )

    print(f"    gentle pause: {delay:.1f} seconds")
    time.sleep(delay)
    return delay


def cache_key_for_url(url: str) -> str:
    """
    Turn a Wikidot URL into a filesystem-safe cache filename.
    """
    slug = url.rstrip("/").split("/")[-1]
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", slug)
    return slug + ".html"


def get_cache_path(cache_folder: Path, url: str) -> Path:
    return cache_folder / cache_key_for_url(url)


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept-Language": "en-US,en;q=0.9",
        }
    )
    return session


def fetch_soup(
    session: requests.Session,
    url: str,
    cache_folder: Optional[Path] = None,
    use_cache: bool = False,
) -> BeautifulSoup:
    """
    Fetch a page with conservative retry/backoff behavior.

    Item pages can be cached locally. The master page is intentionally fetched
    fresh once per run.
    """
    cache_path = None

    if use_cache and cache_folder is not None:
        cache_folder.mkdir(parents=True, exist_ok=True)
        cache_path = get_cache_path(cache_folder, url)

        if cache_path.is_file():
            print("    cache hit:", cache_path.name)
            return BeautifulSoup(
                cache_path.read_text(
                    encoding="utf-8",
                    errors="replace",
                ),
                "html.parser",
            )

    last_error = None

    for attempt in range(MAX_RETRIES_PER_PAGE):
        try:
            response = session.get(
                url,
                timeout=REQUEST_TIMEOUT_SECONDS,
            )

            status = response.status_code

            if status == 429:
                retry_after = response.headers.get("Retry-After")

                print()
                print("Wikidot returned HTTP 429 (Too Many Requests).")

                if retry_after:
                    print("Server Retry-After:", retry_after)

                if STOP_ON_HTTP_429:
                    raise StopGentleRun(
                        "Wikidot asked us to slow down. The run was stopped."
                    )

            if status == 403:
                last_error = requests.HTTPError(
                    f"HTTP 403 for {url}"
                )

                if attempt + 1 >= MAX_RETRIES_PER_PAGE:
                    if STOP_ON_REPEATED_SERVER_ERRORS:
                        raise StopGentleRun(
                            "Repeated HTTP 403 responses. "
                            "Stopping instead of repeatedly requesting the site."
                        )

            elif 500 <= status <= 599:
                last_error = requests.HTTPError(
                    f"HTTP {status} for {url}"
                )

                if attempt + 1 >= MAX_RETRIES_PER_PAGE:
                    if STOP_ON_REPEATED_SERVER_ERRORS:
                        raise StopGentleRun(
                            f"Repeated HTTP {status} server errors. "
                            "Stopping the run."
                        )

            else:
                response.raise_for_status()

                text = response.text

                if cache_path is not None:
                    cache_path.write_text(
                        text,
                        encoding="utf-8",
                    )

                return BeautifulSoup(text, "html.parser")

        except StopGentleRun:
            raise

        except requests.RequestException as error:
            last_error = error

        # Back off before retrying.
        if attempt + 1 < MAX_RETRIES_PER_PAGE:
            backoff = BACKOFF_SECONDS[
                min(attempt, len(BACKOFF_SECONDS) - 1)
            ]

            print(
                f"    temporary request problem; "
                f"backing off {backoff} seconds before retry "
                f"{attempt + 2}/{MAX_RETRIES_PER_PAGE}"
            )
            time.sleep(backoff)

    if last_error:
        raise last_error

    raise RuntimeError(f"Could not fetch {url}")


def collect_item_links(session: requests.Session) -> Dict[str, Tuple[str, str]]:
    """
    Read the master Magic Items page and build:

        normalized item name -> (display name, absolute URL)

    Wikidot item pages use paths such as:
        /wondrous-items:bag-of-holding
    """
    print("Reading master Magic Items page...")
    soup = fetch_soup(session, MASTER_URL)

    links: Dict[str, Tuple[str, str]] = {}

    # Prefer the page-content section so navigation links are ignored.
    content = soup.select_one("#page-content") or soup

    for anchor in content.find_all("a", href=True):
        display_name = anchor.get_text(" ", strip=True)
        href = anchor.get("href", "").strip()

        if not display_name or not href:
            continue

        absolute_url = urljoin(MASTER_URL, href)

        # Item links on this page are in the wondrous-items namespace.
        if "/wondrous-items:" not in absolute_url:
            continue

        key = normalize_name(display_name)

        if key and key not in links:
            links[key] = (display_name, absolute_url)

    print(f"Master-page item links found: {len(links)}")
    return links


def find_header_columns(ws) -> Dict[str, int]:
    """
    Return header text -> 1-based column number from row 1.
    """
    headers = {}

    for cell in ws[1]:
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = cell.column

    return headers


def ensure_url_column(ws, headers: Dict[str, int]) -> int:
    """
    Reuse Wikidot URL if present; otherwise add it after the current final column.
    """
    if URL_HEADER in headers:
        return headers[URL_HEADER]

    new_col = ws.max_column + 1
    header_cell = ws.cell(row=1, column=new_col, value=URL_HEADER)

    # Copy the look of the previous header where possible.
    if new_col > 1:
        previous = ws.cell(row=1, column=new_col - 1)
        if previous.has_style:
            header_cell._style = copy(previous._style)
        if previous.font:
            header_cell.font = copy(previous.font)
        if previous.fill:
            header_cell.fill = copy(previous.fill)
        if previous.border:
            header_cell.border = copy(previous.border)
        if previous.alignment:
            header_cell.alignment = copy(previous.alignment)
        if previous.number_format:
            header_cell.number_format = previous.number_format
        if previous.protection:
            header_cell.protection = copy(previous.protection)

    ws.column_dimensions[header_cell.column_letter].width = 48
    return new_col


def exact_or_fuzzy_match(
    item_name: str,
    links: Dict[str, Tuple[str, str]],
) -> Tuple[Optional[Tuple[str, str]], str, float]:
    """
    Return:
        ((master display name, URL) or None, match_type, similarity)

    Fuzzy matching is deliberately conservative.
    """
    wanted = normalize_name(item_name)

    if wanted in links:
        return links[wanted], "exact", 1.0

    best_key = None
    best_score = 0.0

    for candidate in links.keys():
        score = SequenceMatcher(None, wanted, candidate).ratio()

        if score > best_score:
            best_score = score
            best_key = candidate

    if best_key and best_score >= FUZZY_MATCH_THRESHOLD:
        return links[best_key], "fuzzy", best_score

    return None, "unmatched", best_score


def clean_page_content(soup: BeautifulSoup) -> BeautifulSoup:
    """
    Remove Wikidot editing/navigation/tag controls from the main item content.
    """
    content = soup.select_one("#page-content")

    if content is None:
        raise ValueError("Could not find #page-content on item page.")

    # Work on a parsed copy so we do not depend on the original tree later.
    content = BeautifulSoup(str(content), "html.parser")

    selectors_to_remove = [
        "script",
        "style",
        ".page-tags",
        "#page-options-bottom",
        "#page-options-bottom-2",
        ".page-watch-options",
        ".rate-box-with-credit-button",
        ".page-rate-widget-box",
        ".printuser",
    ]

    for selector in selectors_to_remove:
        for node in content.select(selector):
            node.decompose()

    return content


def looks_like_item_type_line(text: str) -> bool:
    """
    Detect the line immediately below Source:, e.g.
        Wondrous item, uncommon
        Weapon (longsword), legendary (requires attunement)
    """
    lower = text.lower()

    rarity_words = (
        "common",
        "uncommon",
        "rare",
        "very rare",
        "legendary",
        "artifact",
        "unique",
    )

    item_words = (
        "wondrous item",
        "weapon",
        "armor",
        "ring",
        "rod",
        "staff",
        "wand",
        "potion",
        "scroll",
        "ammunition",
    )

    return (
        any(word in lower for word in rarity_words)
        and any(word in lower for word in item_words)
    )


def extract_description(session: requests.Session, url: str, cache_folder: Path) -> str:
    """
    Extract the body description while omitting:
      - page title
      - Source:
      - item type/rarity line
      - page tags
      - Wikidot edit/navigation controls

    Tables/lists/headings are retained as readable text.
    """
    soup = fetch_soup(session, url, cache_folder=cache_folder, use_cache=True)
    content = clean_page_content(soup)

    # Convert useful block structure into line breaks before get_text().
    for br in content.find_all("br"):
        br.replace_with("\n")

    for tag_name in ("p", "div", "li", "tr", "h1", "h2", "h3", "h4", "h5", "h6"):
        for node in content.find_all(tag_name):
            node.append("\n")

    raw_text = html.unescape(content.get_text("\n", strip=False))

    lines = []
    for raw_line in raw_text.splitlines():
        line = re.sub(r"[ \t]+", " ", raw_line).strip()
        if line:
            lines.append(line)

    if not lines:
        raise ValueError("Item page did not contain readable text.")

    # Start immediately after Source: and the item type/rarity line.
    start_index = 0

    for index, line in enumerate(lines):
        if line.lower().startswith("source:"):
            start_index = index + 1
            break

    while start_index < len(lines) and looks_like_item_type_line(lines[start_index]):
        start_index += 1

    # Some pages may have one blank-equivalent/meta line between source and body.
    while start_index < len(lines):
        low = lines[start_index].lower()

        if low.startswith("source:") or looks_like_item_type_line(lines[start_index]):
            start_index += 1
            continue

        break

    body_lines = []

    stop_phrases = (
        "click here to edit contents of this page",
        "click here to toggle editing",
        "append content without editing",
        "check out how this page has evolved",
        "if you want to discuss contents of this page",
        "view and manage file attachments",
        "a few useful tools to manage this site",
        "see pages that link to and include this page",
        "change the name",
        "view wiki source",
        "view/set parent page",
        "notify administrators",
        "something does not work as expected",
        "general wikidot.com documentation",
        "wikidot.com terms of service",
        "wikidot.com privacy policy",
    )

    for line in lines[start_index:]:
        low = line.lower()

        if any(low.startswith(phrase) for phrase in stop_phrases):
            break

        body_lines.append(line)

    # Collapse accidental repeated adjacent lines caused by nested div/p tags.
    deduped = []
    for line in body_lines:
        if deduped and line == deduped[-1]:
            continue
        deduped.append(line)

    description = "\n".join(deduped).strip()

    if not description:
        raise ValueError("Could not isolate the description body.")

    return description


def make_backup(input_path: Path) -> Path:
    backup = input_path.with_name(
        input_path.stem + " - BEFORE WIKIDOT AUTOFILL" + input_path.suffix
    )

    if not backup.exists():
        shutil.copy2(input_path, backup)

    return backup


def save_checkpoint(wb, output_path: Path) -> None:
    temp_path = output_path.with_suffix(".tmp.xlsx")
    wb.save(temp_path)
    os.replace(temp_path, output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Fill missing D&D magic-item descriptions from the D&D 5e Wikidot master item list."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default=DEFAULT_WORKBOOK,
        help=f'Workbook path (default: "{DEFAULT_WORKBOOK}")',
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Check matching only. Do not fetch descriptions or save workbook changes.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help=(
            "Process at most this many missing descriptions. "
            f"Default cap is {DEFAULT_MAX_ITEMS_PER_RUN}."
        ),
    )
    args = parser.parse_args()

    input_path = Path(args.workbook).expanduser().resolve()

    if not input_path.is_file():
        print()
        print("ERROR: Workbook not found:")
        print(input_path)
        print()
        print("Put this Python file beside DND info.xlsx, or pass the workbook path.")
        return 1

    output_path = input_path.with_name(
        input_path.stem + " - AUTOFILLED" + input_path.suffix
    )
    report_path = input_path.with_name("dnd_wikidot_autofill_report.txt")
    cache_folder = input_path.parent / CACHE_FOLDER_NAME

    print()
    print("=" * 72)
    print("D&D WIKIDOT MAGIC ITEM AUTOFILLER")
    print("=" * 72)
    print("Input: ", input_path)
    print("Output:", output_path)
    print("Sheet: ", TARGET_SHEET)
    print("Mode:  ", "DRY RUN" if args.dry_run else "WRITE DESCRIPTIONS")
    print(
        "Gentle delay:",
        f"{MIN_REQUEST_DELAY_SECONDS:.0f}-{MAX_REQUEST_DELAY_SECONDS:.0f} seconds between uncached item requests",
    )
    print("Default run cap:", DEFAULT_MAX_ITEMS_PER_RUN, "items")
    print("Cache folder:", cache_folder)
    print()

    session = make_session()

    try:
        item_links = collect_item_links(session)
    except Exception as error:
        print("ERROR: Could not read the Wikidot master page:")
        print(error)
        return 1

    print("Opening workbook...")

    try:
        # Work on the original content but write to a separate output workbook.
        wb = load_workbook(input_path)
    except Exception as error:
        print("ERROR: Could not open workbook:")
        print(error)
        return 1

    if TARGET_SHEET not in wb.sheetnames:
        print(f'ERROR: Sheet "{TARGET_SHEET}" was not found.')
        print("Available sheets:")
        for name in wb.sheetnames:
            print(" -", name)
        return 1

    ws = wb[TARGET_SHEET]
    headers = find_header_columns(ws)

    if ITEM_NAME_HEADER not in headers:
        print(f'ERROR: Could not find "{ITEM_NAME_HEADER}" in row 1.')
        return 1

    if DESCRIPTION_HEADER not in headers:
        print(f'ERROR: Could not find "{DESCRIPTION_HEADER}" in row 1.')
        return 1

    item_col = headers[ITEM_NAME_HEADER]
    description_col = headers[DESCRIPTION_HEADER]
    url_col = ensure_url_column(ws, headers)

    # First match every item in the workbook to the master page.
    # This gives the sheet a useful clickable source URL even for rows whose
    # descriptions were already completed manually.
    all_row_matches = {}
    all_unmatched_names = []

    for row in range(2, ws.max_row + 1):
        item_name = ws.cell(row=row, column=item_col).value

        if not item_name:
            continue

        item_name = str(item_name).strip()
        match, match_type, score = exact_or_fuzzy_match(item_name, item_links)

        if match is None:
            all_unmatched_names.append((row, item_name, score))
            continue

        master_name, url = match
        all_row_matches[row] = (master_name, url, match_type, score)

        if not args.dry_run:
            url_cell = ws.cell(row=row, column=url_col)
            url_cell.value = url
            url_cell.hyperlink = url
            url_cell.style = "Hyperlink"

    missing_rows = []

    for row in range(2, ws.max_row + 1):
        item_name = ws.cell(row=row, column=item_col).value
        description = ws.cell(row=row, column=description_col).value

        if not item_name:
            continue

        if description is None or str(description).strip() == "":
            missing_rows.append((row, str(item_name).strip()))

    print(f"Rows in sheet:               {ws.max_row - 1}")
    print(f"Item URLs matched:           {len(all_row_matches)}")
    print(f"Item URLs unmatched:         {len(all_unmatched_names)}")
    print(f"Missing descriptions found: {len(missing_rows)}")
    print()

    requested_limit = (
        max(0, args.limit)
        if args.limit is not None
        else DEFAULT_MAX_ITEMS_PER_RUN
    )

    # Gentle mode has a true hard ceiling. Even --limit 500 will not exceed
    # DEFAULT_MAX_ITEMS_PER_RUN unless the source code setting is deliberately
    # changed.
    requested_limit = min(
        requested_limit,
        DEFAULT_MAX_ITEMS_PER_RUN,
    )

    if len(missing_rows) > requested_limit:
        missing_rows = missing_rows[:requested_limit]

    print(
        f"Run cap active: processing at most {len(missing_rows)} missing row(s)."
    )
    print()

    if not args.dry_run:
        backup_path = make_backup(input_path)
        print("Safety backup:", backup_path)
        print()

    filled = 0
    exact_matches = 0
    fuzzy_matches = 0
    unmatched = []
    failed = []

    for number, (row, item_name) in enumerate(missing_rows, start=1):
        cached_match = all_row_matches.get(row)

        prefix = f"[{number}/{len(missing_rows)}] Row {row}: {item_name}"

        if cached_match is None:
            # Recover the best similarity for the report.
            _, _, score = exact_or_fuzzy_match(item_name, item_links)
            print(prefix)
            print(f"    NOT FOUND (best similarity {score:.3f})")
            unmatched.append((row, item_name, score))
            continue

        master_name, url, match_type, score = cached_match

        if match_type == "exact":
            exact_matches += 1
        else:
            fuzzy_matches += 1

        print(prefix)
        if match_type == "fuzzy":
            print(
                f'    FUZZY MATCH {score:.3f}: "{master_name}"'
            )

        if args.dry_run:
            print("    MATCHED:", url)
            continue

        cache_path = get_cache_path(cache_folder, url)
        was_cached = cache_path.is_file()

        try:
            description = extract_description(
                session,
                url,
                cache_folder,
            )

            ws.cell(row=row, column=description_col).value = description

            # Make long descriptions readable in Excel/Sheets.
            desc_cell = ws.cell(row=row, column=description_col)
            desc_cell.alignment = copy(desc_cell.alignment)
            desc_cell.alignment = desc_cell.alignment.copy(
                wrap_text=True,
                vertical="top",
            )

            filled += 1
            print(f"    FILLED ({len(description):,} characters)")

        except StopGentleRun as error:
            print()
            print("=" * 72)
            print("GENTLE MODE STOP")
            print("=" * 72)
            print(error)
            print("Saving current workbook progress before exiting the item loop.")
            failed.append((row, item_name, url, str(error)))

            if not args.dry_run:
                save_checkpoint(wb, output_path)

            break

        except Exception as error:
            print("    ERROR:", error)
            failed.append((row, item_name, url, str(error)))

        # Save incrementally.
        if (
            not args.dry_run
            and number % SAVE_EVERY_N_ITEMS == 0
        ):
            save_checkpoint(wb, output_path)
            print("    checkpoint saved")

        # Cached pages create no Wikidot traffic, so there is no need to
        # wait after a cache hit. New network requests get randomized spacing.
        if not args.dry_run and not was_cached:
            gentle_delay()

    if not args.dry_run:
        save_checkpoint(wb, output_path)

    report_lines = [
        "D&D Wikidot Autofill Report",
        "=" * 72,
        f"Workbook: {input_path}",
        f"Sheet: {TARGET_SHEET}",
        f"Missing rows considered: {len(missing_rows)}",
        f"Descriptions filled: {filled}",
        f"Exact matches: {exact_matches}",
        f"Fuzzy matches: {fuzzy_matches}",
        f"Unmatched: {len(unmatched)}",
        f"Failed downloads/extractions: {len(failed)}",
        f"Gentle delay range: {MIN_REQUEST_DELAY_SECONDS}-{MAX_REQUEST_DELAY_SECONDS} seconds",
        f"Default max items per run: {DEFAULT_MAX_ITEMS_PER_RUN}",
        f"Cache folder: {cache_folder}",
        "",
    ]

    if unmatched:
        report_lines.append("UNMATCHED ITEMS")
        report_lines.append("-" * 72)
        for row, name, score in unmatched:
            report_lines.append(
                f"Row {row}: {name} | best similarity={score:.3f}"
            )
        report_lines.append("")

    if failed:
        report_lines.append("FAILED ITEMS")
        report_lines.append("-" * 72)
        for row, name, url, error in failed:
            report_lines.append(
                f"Row {row}: {name}\nURL: {url}\nError: {error}\n"
            )

    report_path.write_text(
        "\n".join(report_lines),
        encoding="utf-8",
    )

    print()
    print("=" * 72)
    print("DONE")
    print("=" * 72)

    if args.dry_run:
        print("Dry run only; workbook was not changed.")
    else:
        print("Autofilled workbook:")
        print(output_path)

    print()
    print("Report:")
    print(report_path)
    print()
    print(f"Descriptions filled: {filled}")
    print(f"Unmatched:           {len(unmatched)}")
    print(f"Failed:              {len(failed)}")
    print()

    if unmatched or failed:
        print("Review the report for anything that needs manual attention.")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
