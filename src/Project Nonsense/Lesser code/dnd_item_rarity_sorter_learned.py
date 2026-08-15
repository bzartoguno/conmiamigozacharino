#!/usr/bin/env python3
"""
D&D Item Rarity Sorter - Learned From Canon D&D Items
=====================================================

Purpose
-------
Assign rarity to entries in:
    Copy of Current DND items

using the known-rarity items in:
    DBCannonDndItems for programing

as examples of what Common / Uncommon / Rare / Very Rare / Legendary /
Artifact effects look like.

PRICE IS NEVER READ OR USED.

The classifier uses:
1. Similarity to your canon D&D item descriptions.
2. Explicit mechanical power cues:
   - numerical bonuses
   - advantage/disadvantage
   - resistance/immunity
   - flight/invisibility/teleportation
   - healing/resurrection/death effects
   - summons/companions
   - permanent effects
   - consumable / one-use limitations
   - hireling/service duration and scope
   - town skill-tree control
3. A review sheet for uncertain cases.

The program preserves any rarity you manually set unless --overwrite is used.

Install:
    python -m pip install openpyxl

Normal run:
    python dnd_item_rarity_sorter_learned.py

Dry run:
    python dnd_item_rarity_sorter_learned.py --dry-run

Recalculate existing rarities too:
    python dnd_item_rarity_sorter_learned.py --overwrite

Specific workbook:
    python dnd_item_rarity_sorter_learned.py "E:\\DND infov2.xlsx"
"""

import argparse
import math
import os
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# WORKBOOK SETTINGS
# ---------------------------------------------------------------------------

TRAINING_SHEET_EXACT = "DBCannonDndItems for programing"
TRAINING_SHEET_CONTAINS = "DBCannonDndItems"
TARGET_SHEET = "Copy of Current DND items"

ITEM_NAME_HEADER = "Item Name"
RARITY_HEADER = "Rarity"
ITEM_TYPE_HEADER = "Item Type"
ATTUNEMENT_HEADER = "Attunment"
DESCRIPTION_HEADER = "Description"

REVIEW_SHEET_NAME = "Rarity Review"
OUTPUT_SUFFIX = " - RARITY SORTED"
BACKUP_SUFFIX = " - BEFORE RARITY SORT"

UNKNOWN_VALUES = {
    "",
    "?",
    "???",
    "????",
    "NA",
    "N/A",
    "NONE",
    "TBD",
}

RARITY_LEVELS = [
    "Common",
    "Uncommon",
    "Rare",
    "Very Rare",
    "Legendary",
    "Artifact",
]

RARITY_NUMBER = {
    rarity: index + 1
    for index, rarity in enumerate(RARITY_LEVELS)
}

NUMBER_TO_RARITY = {
    value: key
    for key, value in RARITY_NUMBER.items()
}


# ---------------------------------------------------------------------------
# USER-EDITABLE EXACT OVERRIDES
# ---------------------------------------------------------------------------
#
# If you later decide an item should ALWAYS have one specific rarity,
# put it here:
#
# RARITY_OVERRIDES = {
#     "Charon's Token": "Rare",
# }
#
RARITY_OVERRIDES: Dict[str, str] = {
    # These entries explicitly describe the rarity level they deal with.
    "Sell a Trash Item": "Common",
    "Sell a Common Item": "Common",
    "Sell an Uncommon Item": "Uncommon",
    "Sell a Rare Item": "Rare",
    "Sell a Very Rare Item": "Very Rare",
    "Sell a Legendary Item": "Legendary",
    "Sell an Artifact Item": "Artifact",
}


# ---------------------------------------------------------------------------
# TEXT MODEL
# ---------------------------------------------------------------------------

STOPWORDS = set(
    """
    a an and are as at be been being but by can could did do does for from had
    has have he her hers him his how i if in into is it its may me more most my
    no not of on one only or our out over she so some than that the their them
    then there these they this to too up use used uses using was we were what
    when where which while who will with you your

    action attack creature creatures item items magical magic target targets
    turn make makes made effect effects number roll rolls gain gains gained
    wearing wielding wielder
    """.split()
)


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalized(value) -> str:
    text = clean_text(value).lower()
    text = (
        text.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("–", "-")
        .replace("—", "-")
    )
    return re.sub(r"\s+", " ", text).strip()


def phrase_present(text: str, phrase: str) -> bool:
    text = normalized(text)
    phrase = normalized(phrase)

    if not phrase:
        return False

    if " " in phrase:
        return phrase in text

    return re.search(
        rf"\b{re.escape(phrase)}\b",
        text,
    ) is not None


def tokenise(value) -> List[str]:
    text = normalized(value)

    words = re.findall(
        r"[a-z][a-z0-9']+",
        text,
    )

    tokens: List[str] = []

    for word in words:
        if word in STOPWORDS:
            continue

        if len(word) <= 2:
            continue

        if word.endswith("'s"):
            word = word[:-2]

        # Gentle stemming so "teleports" and "teleporting" resemble teleport.
        for suffix in ("ing", "ed", "es", "s"):
            if (
                len(word) > 6
                and word.endswith(suffix)
            ):
                word = word[:-len(suffix)]
                break

        if word and word not in STOPWORDS:
            tokens.append(word)

    return tokens


def weighted_document(
    name: str,
    item_type: str,
    description: str,
) -> List[str]:
    return (
        tokenise(name) * 4
        + tokenise(item_type) * 2
        + tokenise(description)
    )


def rarity_is_unknown(value) -> bool:
    return normalized(value).upper() in UNKNOWN_VALUES


# ---------------------------------------------------------------------------
# TRAINING MODEL
# ---------------------------------------------------------------------------

class RarityExample:
    def __init__(
        self,
        row_number: int,
        name: str,
        rarity: str,
        item_type: str,
        attunement: bool,
        description: str,
    ):
        self.row_number = row_number
        self.name = name
        self.rarity = rarity
        self.item_type = item_type
        self.attunement = attunement
        self.description = description
        self.tokens = weighted_document(
            name,
            item_type,
            description,
        )
        self.vector: Dict[str, float] = {}


class RarityModel:
    def __init__(
        self,
        examples: List[RarityExample],
    ):
        self.examples = examples
        self.idf: Dict[str, float] = {}
        self.centroids: Dict[str, Dict[str, float]] = {}
        self.examples_by_rarity: Dict[
            str,
            List[RarityExample],
        ] = defaultdict(list)

        for example in examples:
            self.examples_by_rarity[
                example.rarity
            ].append(example)

        self._fit()

    def _fit(self) -> None:
        document_frequency: Counter = Counter()

        for example in self.examples:
            document_frequency.update(
                set(example.tokens)
            )

        document_count = max(
            1,
            len(self.examples),
        )

        self.idf = {
            token: (
                math.log(
                    (document_count + 1)
                    / (frequency + 1)
                )
                + 1.0
            )
            for token, frequency
            in document_frequency.items()
        }

        for example in self.examples:
            example.vector = self.vectorise(
                example.tokens
            )

        for rarity, examples in self.examples_by_rarity.items():
            accumulator: Dict[
                str,
                float,
            ] = defaultdict(float)

            for example in examples:
                for token, value in example.vector.items():
                    accumulator[token] += value

            count = max(
                1,
                len(examples),
            )

            centroid = {
                token: value / count
                for token, value
                in accumulator.items()
            }

            self.centroids[rarity] = (
                self.normalise_vector(
                    centroid
                )
            )

    def vectorise(
        self,
        tokens: List[str],
    ) -> Dict[str, float]:
        counts = Counter(tokens)
        vector: Dict[str, float] = {}

        for token, count in counts.items():
            idf = self.idf.get(
                token,
                1.0,
            )

            vector[token] = (
                (1.0 + math.log(count))
                * idf
            )

        return self.normalise_vector(
            vector
        )

    @staticmethod
    def normalise_vector(
        vector: Dict[str, float],
    ) -> Dict[str, float]:
        magnitude = math.sqrt(
            sum(
                value * value
                for value in vector.values()
            )
        )

        if magnitude <= 0:
            return {}

        return {
            token: value / magnitude
            for token, value
            in vector.items()
        }

    @staticmethod
    def cosine(
        left: Dict[str, float],
        right: Dict[str, float],
    ) -> float:
        if len(left) > len(right):
            left, right = right, left

        return sum(
            value
            * right.get(token, 0.0)
            for token, value
            in left.items()
        )

    def learned_estimate(
        self,
        name: str,
        item_type: str,
        description: str,
    ) -> Tuple[
        float,
        List[Tuple[float, RarityExample]],
        Dict[str, float],
    ]:
        target_vector = self.vectorise(
            weighted_document(
                name,
                item_type,
                description,
            )
        )

        all_neighbors: List[
            Tuple[float, RarityExample]
        ] = []

        rarity_scores: Dict[
            str,
            float,
        ] = defaultdict(float)

        for example in self.examples:
            similarity = self.cosine(
                target_vector,
                example.vector,
            )

            if similarity > 0:
                all_neighbors.append(
                    (
                        similarity,
                        example,
                    )
                )

        all_neighbors.sort(
            key=lambda pair: pair[0],
            reverse=True,
        )

        top_neighbors = all_neighbors[:12]

        numerator = 0.0
        denominator = 0.0

        for similarity, example in top_neighbors:
            # Squaring rewards genuinely close examples over vague overlap.
            weight = similarity * similarity

            numerator += (
                weight
                * RARITY_NUMBER[
                    example.rarity
                ]
            )

            denominator += weight

            rarity_scores[
                example.rarity
            ] += weight

        if denominator > 0:
            neighbor_estimate = (
                numerator
                / denominator
            )
        else:
            neighbor_estimate = 2.0

        # Centroid estimate is weaker but makes the result less fragile when
        # a single weird published item happens to have similar wording.
        centroid_numerator = 0.0
        centroid_denominator = 0.0

        for rarity in RARITY_LEVELS:
            similarity = self.cosine(
                target_vector,
                self.centroids.get(
                    rarity,
                    {},
                ),
            )

            if similarity <= 0:
                continue

            centroid_numerator += (
                similarity
                * RARITY_NUMBER[rarity]
            )
            centroid_denominator += similarity

        if centroid_denominator > 0:
            centroid_estimate = (
                centroid_numerator
                / centroid_denominator
            )
        else:
            centroid_estimate = neighbor_estimate

        learned = (
            0.78 * neighbor_estimate
            + 0.22 * centroid_estimate
        )

        return (
            learned,
            top_neighbors[:5],
            dict(rarity_scores),
        )


# ---------------------------------------------------------------------------
# MECHANICAL POWER ANALYSIS
# ---------------------------------------------------------------------------

def extract_numbers(
    text: str,
) -> List[int]:
    return [
        int(value)
        for value in re.findall(
            r"(?<!\w)\+?(\d+)(?!\w)",
            text,
        )
    ]


def detect_bonus_minimum(
    description: str,
) -> Tuple[float, List[str]]:
    """
    Translate obvious +1 / +2 / +3 style mechanical bonuses into minimum
    rarity pressure.
    """
    text = normalized(description)
    floor = 1.0
    reasons: List[str] = []

    patterns = [
        (
            r"\+3\s+(?:bonus\s+)?(?:to\s+)?(?:ac|attack|damage|saving|save|spell)",
            4.6,
            "+3-class mechanical bonus",
        ),
        (
            r"\+2\s+(?:bonus\s+)?(?:to\s+)?(?:ac|attack|damage|saving|save|spell)",
            3.5,
            "+2-class mechanical bonus",
        ),
        (
            r"\+1\s+(?:bonus\s+)?(?:to\s+)?(?:ac|attack|damage|saving|save|spell)",
            2.3,
            "+1-class mechanical bonus",
        ),
    ]

    for pattern, minimum, reason in patterns:
        if re.search(
            pattern,
            text,
        ):
            floor = max(
                floor,
                minimum,
            )
            reasons.append(
                reason
            )

    return floor, reasons


def mechanical_estimate(
    name: str,
    item_type: str,
    attunement: bool,
    description: str,
) -> Tuple[
    float,
    float,
    List[str],
]:
    """
    Return:
        power_estimate
        minimum_tier_pressure
        reasons

    Values are on the 1-6 rarity scale.
    """
    name_text = normalized(name)
    desc = normalized(description)
    combined = name_text + " " + desc

    # Mundane / flavor items begin at Common.
    score = 1.15
    floor = 1.0
    reasons: List[str] = []

    # Attunement is only a small signal. Plenty of Common items require it.
    if attunement:
        score += 0.18
        reasons.append(
            "requires attunement"
        )

    # -------------------------------------------------------------------
    # Services / hirelings / scope.
    # -------------------------------------------------------------------

    if phrase_present(
        name_text,
        "hire",
    ):
        score = max(
            score,
            1.20,
        )
        reasons.append(
            "hireling/service entry"
        )

        if any(
            phrase_present(
                combined,
                phrase,
            )
            for phrase in (
                "one month",
                "for a month",
                "four weeks",
            )
        ):
            score += 0.45
            reasons.append(
                "month-long service"
            )

        if any(
            phrase_present(
                combined,
                phrase,
            )
            for phrase in (
                "one year",
                "for one year",
            )
        ):
            score += 1.40
            floor = max(
                floor,
                3.0,
            )
            reasons.append(
                "year-long service"
            )

        if any(
            phrase_present(
                combined,
                phrase,
            )
            for phrase in (
                "capture an npc",
                "kill any monster",
                "bounty hunter",
            )
        ):
            floor = max(
                floor,
                3.0,
            )
            score += 1.4
            reasons.append(
                "service can replace meaningful adventuring work"
            )

    if phrase_present(
        combined,
        "town skill tree",
    ):
        score += 2.0
        floor = max(
            floor,
            4.0,
        )
        reasons.append(
            "changes town skill-tree systems"
        )

    if phrase_present(
        combined,
        "take control of",
    ) and phrase_present(
        combined,
        "branches",
    ):
        score += 1.2
        floor = max(
            floor,
            4.0,
        )
        reasons.append(
            "controls multiple town branches"
        )

    # -------------------------------------------------------------------
    # Consumable / frequency.
    # -------------------------------------------------------------------

    one_use = any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "one-time use",
            "one time use",
            "single use",
            "consumed after",
            "is consumed",
            "next check",
            "next attack",
            "next saving throw",
        )
    )

    if one_use:
        score -= 0.30
        reasons.append(
            "limited/one-use effect"
        )

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "at will",
            "whenever you want",
            "without expending",
            "unlimited",
        )
    ):
        score += 0.65
        reasons.append(
            "repeatable/at-will effect"
        )

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "once per day",
            "once a day",
            "daily at dawn",
        )
    ):
        score += 0.15
        reasons.append(
            "recharges daily"
        )

    # -------------------------------------------------------------------
    # Advantage / broad roll manipulation.
    # -------------------------------------------------------------------

    if phrase_present(
        desc,
        "advantage on your next",
    ):
        score += 0.65
        floor = max(
            floor,
            1.6,
        )
        reasons.append(
            "one-shot advantage"
        )

    elif phrase_present(
        desc,
        "advantage on",
    ):
        score += 1.05
        floor = max(
            floor,
            2.0,
        )
        reasons.append(
            "repeatable advantage"
        )

    if phrase_present(
        desc,
        "disadvantage on",
    ):
        score += 0.85
        floor = max(
            floor,
            2.0,
        )
        reasons.append(
            "imposes disadvantage"
        )

    # -------------------------------------------------------------------
    # Numeric bonus floors.
    # -------------------------------------------------------------------

    bonus_floor, bonus_reasons = (
        detect_bonus_minimum(
            description
        )
    )

    floor = max(
        floor,
        bonus_floor,
    )
    reasons.extend(
        bonus_reasons
    )

    # -------------------------------------------------------------------
    # Defense.
    # -------------------------------------------------------------------

    if phrase_present(
        desc,
        "resistance to",
    ):
        score += 1.25
        floor = max(
            floor,
            3.0,
        )
        reasons.append(
            "damage resistance"
        )

    if phrase_present(
        desc,
        "immunity to",
    ) or phrase_present(
        desc,
        "immune to",
    ):
        score += 2.0
        floor = max(
            floor,
            4.0,
        )
        reasons.append(
            "damage/condition immunity"
        )

    # -------------------------------------------------------------------
    # Major utility powers.
    # -------------------------------------------------------------------

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "invisible",
            "invisibility",
        )
    ):
        score += 1.4
        floor = max(
            floor,
            3.0,
        )
        reasons.append(
            "invisibility"
        )

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "flying speed",
            "can fly",
            "gain a fly",
            "broom of flying",
        )
    ):
        score += 1.35
        floor = max(
            floor,
            3.0,
        )
        reasons.append(
            "flight"
        )

    if phrase_present(
        combined,
        "teleport",
    ):
        score += 1.55
        floor = max(
            floor,
            3.0,
        )
        reasons.append(
            "teleportation"
        )

        if any(
            phrase_present(
                combined,
                phrase,
            )
            for phrase in (
                "anywhere",
                "any location",
                "another plane",
                "plane",
            )
        ):
            score += 1.0
            floor = max(
                floor,
                4.0,
            )
            reasons.append(
                "large-scale/planar teleportation"
            )

    # -------------------------------------------------------------------
    # Death / resurrection.
    # -------------------------------------------------------------------

    if phrase_present(
        combined,
        "revivify",
    ):
        score += 1.8
        floor = max(
            floor,
            3.0,
        )
        reasons.append(
            "revivify-level effect"
        )

    if phrase_present(
        combined,
        "raise dead",
    ):
        score += 2.4
        floor = max(
            floor,
            4.0,
        )
        reasons.append(
            "raise-dead-level effect"
        )

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "resurrection",
            "restore to life",
            "return to life",
        )
    ):
        score += 2.8
        floor = max(
            floor,
            4.5,
        )
        reasons.append(
            "resurrection effect"
        )

    if phrase_present(
        combined,
        "true resurrection",
    ):
        score += 1.5
        floor = max(
            floor,
            5.0,
        )
        reasons.append(
            "true-resurrection-scale effect"
        )

    if phrase_present(
        combined,
        "wish",
    ):
        score += 4.0
        floor = max(
            floor,
            6.0,
        )
        reasons.append(
            "wish-scale effect"
        )

    # -------------------------------------------------------------------
    # Summons / companions.
    # -------------------------------------------------------------------

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "summon a creature",
            "summons a creature",
            "summon monster",
            "creates a creature",
        )
    ):
        score += 1.2
        floor = max(
            floor,
            2.5,
        )
        reasons.append(
            "summons/creates a creature"
        )

    # -------------------------------------------------------------------
    # Permanent character or campaign effects.
    # -------------------------------------------------------------------

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "permanently",
            "permanent bonus",
            "permanent effect",
        )
    ):
        score += 1.2
        floor = max(
            floor,
            3.0,
        )
        reasons.append(
            "permanent effect"
        )

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in (
            "extra action",
            "additional action",
            "take another turn",
        )
    ):
        score += 2.3
        floor = max(
            floor,
            4.0,
        )
        reasons.append(
            "action-economy breaking effect"
        )

    # -------------------------------------------------------------------
    # Explicitly mundane / flavor entries.
    # -------------------------------------------------------------------

    flavor_markers = (
        "children's book",
        "novel",
        "romance",
        "simple reusable",
        "basic cooking kit",
        "ordinary",
        "mundane",
        "cosmetic only",
        "no mechanical effect",
    )

    if any(
        phrase_present(
            combined,
            phrase,
        )
        for phrase in flavor_markers
    ):
        score = min(
            score,
            1.45,
        )
        reasons.append(
            "mostly mundane/flavor"
        )

    score = max(
        1.0,
        min(
            6.0,
            score,
        ),
    )

    floor = max(
        1.0,
        min(
            6.0,
            floor,
        ),
    )

    return (
        score,
        floor,
        reasons,
    )


# ---------------------------------------------------------------------------
# COMBINE LEARNED + MECHANICAL
# ---------------------------------------------------------------------------

def classify_rarity(
    model: RarityModel,
    item_name: str,
    item_type: str,
    attunement: bool,
    description: str,
) -> Tuple[
    str,
    float,
    str,
    List[str],
    List[str],
]:
    normalized_name = normalized(
        item_name
    )

    for override_name, rarity in RARITY_OVERRIDES.items():
        if normalized(
            override_name
        ) == normalized_name:
            return (
                rarity,
                float(
                    RARITY_NUMBER[
                        rarity
                    ]
                ),
                "High",
                [
                    "Exact user-editable rarity override"
                ],
                [],
            )

    (
        learned,
        neighbors,
        rarity_similarity,
    ) = model.learned_estimate(
        item_name,
        item_type,
        description,
    )

    (
        mechanics,
        minimum_floor,
        reasons,
    ) = mechanical_estimate(
        item_name,
        item_type,
        attunement,
        description,
    )

    # For normal magic items, published examples are very useful.
    # For services / mundane shop offerings, mechanics and scope should matter
    # more because official D&D magic-item text is a weaker analogy.
    is_service_like = any(
        phrase_present(
            item_name,
            phrase,
        )
        for phrase in (
            "hire",
            "contact a",
            "sell a",
            "mini game",
            "custom",
            "inspect",
        )
    )

    if is_service_like:
        raw = (
            0.28 * learned
            + 0.72 * mechanics
        )
    else:
        raw = (
            0.58 * learned
            + 0.42 * mechanics
        )

    raw = max(
        raw,
        minimum_floor,
    )

    # Slightly conservative rounding. We don't want every halfway case to
    # inflate upward.
    rounded_number = int(
        math.floor(
            raw + 0.42
        )
    )

    rounded_number = max(
        1,
        min(
            6,
            rounded_number,
        ),
    )

    rarity = NUMBER_TO_RARITY[
        rounded_number
    ]

    distance_to_boundary = abs(
        raw - round(raw)
    )

    top_neighbor_similarity = (
        neighbors[0][0]
        if neighbors
        else 0.0
    )

    # Confidence:
    # - strong explicit mechanics or strong official similarity = High
    # - close-to-tier-boundary and weak similarity = Review
    if (
        minimum_floor >= rounded_number - 0.10
        and minimum_floor > 1.0
    ):
        confidence = "High"

    elif (
        top_neighbor_similarity >= 0.35
        and distance_to_boundary <= 0.30
    ):
        confidence = "High"

    elif (
        top_neighbor_similarity >= 0.18
        or len(reasons) >= 2
    ):
        confidence = "Medium"

    else:
        confidence = "Review"

    neighbor_strings: List[str] = []

    for similarity, example in neighbors[:5]:
        neighbor_strings.append(
            f"{example.name} — {example.rarity} "
            f"(canon row {example.row_number}, similarity {similarity:.3f})"
        )

    reason_strings = [
        f"learned estimate from canon items: {learned:.2f}",
        f"mechanical/scope estimate: {mechanics:.2f}",
        f"minimum power floor: {minimum_floor:.2f}",
        f"combined rarity score: {raw:.2f}",
    ]

    reason_strings.extend(
        reasons[:8]
    )

    return (
        rarity,
        round(
            raw,
            3,
        ),
        confidence,
        reason_strings,
        neighbor_strings,
    )


# ---------------------------------------------------------------------------
# WORKBOOK HELPERS
# ---------------------------------------------------------------------------

def script_folder() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


def find_default_workbook() -> Optional[Path]:
    folders: List[Path] = []

    for folder in (
        script_folder(),
        Path.cwd(),
    ):
        if folder not in folders:
            folders.append(
                folder
            )

    preferred = (
        "DND infov2.xlsx",
        "DND info(2).xlsx",
        "DND info.xlsx",
        "DND info(1).xlsx",
    )

    for folder in folders:
        for filename in preferred:
            candidate = (
                folder
                / filename
            )

            if candidate.is_file():
                return candidate

    candidates: List[Path] = []

    for folder in folders:
        for candidate in folder.glob(
            "DND info*.xlsx"
        ):
            upper = (
                candidate.name.upper()
            )

            if candidate.name.startswith(
                "~$"
            ):
                continue

            if "RARITY SORTED" in upper:
                continue

            if "BEFORE" in upper:
                continue

            if candidate not in candidates:
                candidates.append(
                    candidate
                )

    if len(candidates) == 1:
        return candidates[0]

    return None


def header_positions(
    ws,
) -> Dict[str, List[int]]:
    positions: Dict[
        str,
        List[int],
    ] = defaultdict(list)

    for cell in ws[1]:
        if cell.value is None:
            continue

        positions[
            clean_text(
                cell.value
            )
        ].append(
            cell.column
        )

    return positions


def first_header_column(
    positions: Dict[str, List[int]],
    header: str,
) -> Optional[int]:
    columns = positions.get(
        header,
        [],
    )

    return (
        min(columns)
        if columns
        else None
    )


def find_training_sheet(
    wb,
):
    if TRAINING_SHEET_EXACT in wb.sheetnames:
        return wb[
            TRAINING_SHEET_EXACT
        ]

    for name in wb.sheetnames:
        if (
            TRAINING_SHEET_CONTAINS.lower()
            in name.lower()
        ):
            return wb[name]

    return None


def read_training_examples(
    wb,
) -> List[RarityExample]:
    ws = find_training_sheet(
        wb
    )

    if ws is None:
        raise ValueError(
            "Could not find the canon D&D item sheet."
        )

    positions = header_positions(
        ws
    )

    name_col = first_header_column(
        positions,
        ITEM_NAME_HEADER,
    )
    rarity_col = first_header_column(
        positions,
        RARITY_HEADER,
    )
    type_col = first_header_column(
        positions,
        ITEM_TYPE_HEADER,
    )
    attunement_col = first_header_column(
        positions,
        ATTUNEMENT_HEADER,
    )
    description_col = first_header_column(
        positions,
        DESCRIPTION_HEADER,
    )

    for header, column in (
        (
            ITEM_NAME_HEADER,
            name_col,
        ),
        (
            RARITY_HEADER,
            rarity_col,
        ),
        (
            ITEM_TYPE_HEADER,
            type_col,
        ),
        (
            ATTUNEMENT_HEADER,
            attunement_col,
        ),
        (
            DESCRIPTION_HEADER,
            description_col,
        ),
    ):
        if column is None:
            raise ValueError(
                f'Canon sheet is missing "{header}".'
            )

    examples: List[
        RarityExample
    ] = []

    for row in range(
        2,
        ws.max_row + 1,
    ):
        rarity = clean_text(
            ws.cell(
                row=row,
                column=rarity_col,
            ).value
        )

        if rarity not in RARITY_NUMBER:
            # Ignore Unique / ???? / any non-standard label.
            continue

        name = clean_text(
            ws.cell(
                row=row,
                column=name_col,
            ).value
        )

        description = clean_text(
            ws.cell(
                row=row,
                column=description_col,
            ).value
        )

        if not name or not description:
            continue

        examples.append(
            RarityExample(
                row_number=row,
                name=name,
                rarity=rarity,
                item_type=clean_text(
                    ws.cell(
                        row=row,
                        column=type_col,
                    ).value
                ),
                attunement=bool(
                    ws.cell(
                        row=row,
                        column=attunement_col,
                    ).value
                ),
                description=description,
            )
        )

    if not examples:
        raise ValueError(
            "No usable known-rarity canon items were found."
        )

    return examples


# ---------------------------------------------------------------------------
# REVIEW SHEET
# ---------------------------------------------------------------------------

def format_review_sheet(
    ws,
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="7030A0",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    widths = {
        "A": 10,
        "B": 34,
        "C": 20,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 80,
        "H": 90,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    for row in ws.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    high_fill = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9",
    )

    medium_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    review_fill = PatternFill(
        fill_type="solid",
        fgColor="FCE4D6",
    )

    for row in range(
        2,
        ws.max_row + 1,
    ):
        confidence = clean_text(
            ws.cell(
                row=row,
                column=6,
            ).value
        )

        fill = (
            high_fill
            if confidence == "High"
            else medium_fill
            if confidence == "Medium"
            else review_fill
        )

        ws.cell(
            row=row,
            column=6,
        ).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        ws.dimensions
    )


# ---------------------------------------------------------------------------
# FILE SAFETY
# ---------------------------------------------------------------------------

def make_backup(
    input_path: Path,
) -> Path:
    backup = input_path.with_name(
        input_path.stem
        + BACKUP_SUFFIX
        + input_path.suffix
    )

    if not backup.exists():
        shutil.copy2(
            input_path,
            backup,
        )

    return backup


def output_path_for(
    input_path: Path,
) -> Path:
    return input_path.with_name(
        input_path.stem
        + OUTPUT_SUFFIX
        + input_path.suffix
    )


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Assign rarity to Copy of Current DND items using known-rarity "
            "canon D&D items plus mechanical power analysis. Price is ignored."
        )
    )

    parser.add_argument(
        "workbook",
        nargs="?",
        default=None,
        help=(
            "Workbook path. If omitted, look beside the script."
        ),
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help=(
            "Analyze without saving a workbook."
        ),
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help=(
            "Recalculate rarities you already filled manually."
        ),
    )

    args = parser.parse_args()

    if args.workbook:
        input_path = Path(
            args.workbook
        ).expanduser().resolve()
    else:
        found = (
            find_default_workbook()
        )

        if found is None:
            print()
            print(
                "ERROR: Could not identify the DND workbook."
            )
            print(
                "Put this script beside DND infov2.xlsx "
                "or pass the workbook path explicitly."
            )
            return 1

        input_path = found.resolve()

    if not input_path.is_file():
        print()
        print(
            "ERROR: Workbook not found:"
        )
        print(
            input_path
        )
        return 1

    print()
    print("=" * 80)
    print("D&D ITEM RARITY SORTER - POWER BASED, NOT PRICE BASED")
    print("=" * 80)
    print(
        "Input:   ",
        input_path,
    )
    print(
        "Training:",
        TRAINING_SHEET_EXACT,
    )
    print(
        "Target:  ",
        TARGET_SHEET,
    )
    print(
        "Mode:    ",
        "DRY RUN"
        if args.dry_run
        else "WRITE RARITIES",
    )
    print(
        "Manual rarities:",
        "OVERWRITE"
        if args.overwrite
        else "PRESERVE",
    )
    print(
        "Price column: IGNORED COMPLETELY"
    )
    print(
        "Internet: NONE"
    )
    print()

    try:
        wb = load_workbook(
            input_path
        )
    except Exception as error:
        print(
            "ERROR: Could not open workbook:"
        )
        print(
            error
        )
        return 1

    try:
        examples = (
            read_training_examples(
                wb
            )
        )
    except Exception as error:
        print(
            "ERROR: Could not build rarity training set:"
        )
        print(
            error
        )
        return 1

    model = RarityModel(
        examples
    )

    training_counts = Counter(
        example.rarity
        for example in examples
    )

    print(
        "Canon training items:",
        len(
            examples
        ),
    )

    for rarity in RARITY_LEVELS:
        print(
            f"  {rarity:<10}",
            training_counts[
                rarity
            ],
        )

    print()

    if TARGET_SHEET not in wb.sheetnames:
        print(
            f'ERROR: Sheet "{TARGET_SHEET}" was not found.'
        )
        print(
            "Available sheets:"
        )

        for name in wb.sheetnames:
            print(
                " -",
                name,
            )

        return 1

    ws = wb[
        TARGET_SHEET
    ]

    positions = header_positions(
        ws
    )

    name_col = first_header_column(
        positions,
        ITEM_NAME_HEADER,
    )
    rarity_col = first_header_column(
        positions,
        RARITY_HEADER,
    )
    type_col = first_header_column(
        positions,
        ITEM_TYPE_HEADER,
    )
    attunement_col = first_header_column(
        positions,
        ATTUNEMENT_HEADER,
    )
    description_col = first_header_column(
        positions,
        DESCRIPTION_HEADER,
    )

    for header, column in (
        (
            ITEM_NAME_HEADER,
            name_col,
        ),
        (
            RARITY_HEADER,
            rarity_col,
        ),
        (
            ITEM_TYPE_HEADER,
            type_col,
        ),
        (
            ATTUNEMENT_HEADER,
            attunement_col,
        ),
        (
            DESCRIPTION_HEADER,
            description_col,
        ),
    ):
        if column is None:
            print(
                f'ERROR: Target sheet is missing "{header}".'
            )
            return 1

    print(
        "Rarity output column:",
        get_column_letter(
            rarity_col
        ),
        f'("{RARITY_HEADER}")',
    )
    print()

    total_items = 0
    assigned = 0
    preserved = 0

    rarity_counts: Counter = Counter()
    confidence_counts: Counter = Counter()

    review_rows: List[
        List
    ] = []

    for excel_row in range(
        2,
        ws.max_row + 1,
    ):
        item_name = clean_text(
            ws.cell(
                row=excel_row,
                column=name_col,
            ).value
        )

        if not item_name:
            continue

        total_items += 1

        current_rarity = ws.cell(
            row=excel_row,
            column=rarity_col,
        ).value

        if (
            not args.overwrite
            and not rarity_is_unknown(
                current_rarity
            )
        ):
            preserved += 1

            rarity_counts[
                clean_text(
                    current_rarity
                )
            ] += 1

            continue

        item_type = clean_text(
            ws.cell(
                row=excel_row,
                column=type_col,
            ).value
        )

        attunement = bool(
            ws.cell(
                row=excel_row,
                column=attunement_col,
            ).value
        )

        description = clean_text(
            ws.cell(
                row=excel_row,
                column=description_col,
            ).value
        )

        (
            rarity,
            score,
            confidence,
            reasons,
            neighbors,
        ) = classify_rarity(
            model=model,
            item_name=item_name,
            item_type=item_type,
            attunement=attunement,
            description=description,
        )

        if not args.dry_run:
            cell = ws.cell(
                row=excel_row,
                column=rarity_col,
            )

            cell.value = rarity

            cell.alignment = Alignment(
                vertical="top",
                horizontal="center",
            )

        assigned += 1
        rarity_counts[
            rarity
        ] += 1
        confidence_counts[
            confidence
        ] += 1

        if confidence != "High":
            review_rows.append(
                [
                    excel_row,
                    item_name,
                    item_type,
                    attunement,
                    rarity,
                    confidence,
                    "\n".join(
                        reasons
                    ),
                    "\n".join(
                        neighbors
                    ),
                ]
            )

    print(
        "Items found:             ",
        total_items,
    )
    print(
        "Rarities generated:      ",
        assigned,
    )
    print(
        "Manual rarities kept:    ",
        preserved,
    )
    print()

    print(
        "CONFIDENCE"
    )
    print(
        "-" * 80
    )

    for label in (
        "High",
        "Medium",
        "Review",
    ):
        print(
            f"{label:<10}",
            confidence_counts[
                label
            ],
        )

    print()
    print(
        "RARITY DISTRIBUTION"
    )
    print(
        "-" * 80
    )

    for rarity in RARITY_LEVELS:
        print(
            f"{rarity:<10}",
            rarity_counts[
                rarity
            ],
        )

    if args.dry_run:
        print()
        print(
            "Dry run complete. No workbook was changed or saved."
        )
        return 0

    if REVIEW_SHEET_NAME in wb.sheetnames:
        del wb[
            REVIEW_SHEET_NAME
        ]

    review_ws = wb.create_sheet(
        REVIEW_SHEET_NAME
    )

    review_ws.append(
        [
            "Excel Row",
            "Item Name",
            "Item Type",
            "Attunement",
            "Suggested Rarity",
            "Confidence",
            "Why It Chose This",
            "Closest Canon D&D Items",
        ]
    )

    for row in review_rows:
        review_ws.append(
            row
        )

    format_review_sheet(
        review_ws
    )

    ws.column_dimensions[
        get_column_letter(
            rarity_col
        )
    ].width = 16

    backup = make_backup(
        input_path
    )

    output_path = (
        output_path_for(
            input_path
        )
    )

    wb.save(
        output_path
    )

    print()
    print("=" * 80)
    print("DONE")
    print("=" * 80)
    print(
        "Original:",
        input_path,
    )
    print()
    print(
        "Safety backup:",
        backup,
    )
    print()
    print(
        "Rarity-sorted workbook:",
        output_path,
    )
    print()
    print(
        f'Review sheet: "{REVIEW_SHEET_NAME}" '
        f"({len(review_rows)} rows)"
    )
    print()
    print(
        "PRICE WAS NOT USED."
    )
    print(
        "If you manually change a rarity in the output workbook, "
        "rerunning this script on that workbook will preserve your choice "
        "unless you use --overwrite."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
